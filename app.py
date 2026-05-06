from flask import Flask, render_template, request, redirect, send_file, session
from zoneinfo import ZoneInfo
ROME = ZoneInfo("Europe/Rome")

def now_rome():
    from datetime import datetime
    return datetime.now(ROME).replace(tzinfo=None)
import os
import psycopg2
from psycopg2.extras import RealDictCursor
from openpyxl import load_workbook
import json

app = Flask(__name__)
app.secret_key = "lc_wine_secret_2026"

DATABASE_URL = os.environ.get("DATABASE_URL")
BASE_DIR     = os.path.dirname(os.path.abspath(__file__))
MODELLI_DIR  = os.path.join(BASE_DIR, "modelli")

# Bottiglie per pedana
BT_PER_PEDANA_2L = 889
BT_PER_PEDANA_1L = 1344


def db():
    return psycopg2.connect(DATABASE_URL, cursor_factory=RealDictCursor)


def init_db():
    conn = db()
    cur = conn.cursor()
    cur.execute("""
        CREATE TABLE IF NOT EXISTS stock (
            id SERIAL PRIMARY KEY, cliente TEXT, prodotto TEXT, qty INTEGER
        )
    """)
    cur.execute("""
        CREATE TABLE IF NOT EXISTS produzione (
            id SERIAL PRIMARY KEY, cliente TEXT, prodotto TEXT,
            qty INTEGER, done INTEGER DEFAULT 0,
            timestamp_done TIMESTAMP
        )
    """)
    # Aggiunge colonna timestamp_done se non esiste (per DB già esistenti)
    cur.execute("""
        ALTER TABLE produzione ADD COLUMN IF NOT EXISTS timestamp_done TIMESTAMP
    """)
    cur.execute("""
        CREATE TABLE IF NOT EXISTS storico (
            id SERIAL PRIMARY KEY, cliente TEXT, prodotto TEXT,
            qty INTEGER, tipo TEXT, data TIMESTAMP DEFAULT NOW()
        )
    """)
    cur.execute("""
        CREATE TABLE IF NOT EXISTS note (
            id SERIAL PRIMARY KEY, testo TEXT, data TIMESTAMP DEFAULT NOW()
        )
    """)
    cur.execute("""
        CREATE TABLE IF NOT EXISTS materie_prime (
            id SERIAL PRIMARY KEY, cliente TEXT, materiale TEXT,
            qty INTEGER DEFAULT 0, soglia_minima INTEGER DEFAULT 0
        )
    """)
    cur.execute("""
        CREATE TABLE IF NOT EXISTS storico_mp (
            id SERIAL PRIMARY KEY, cliente TEXT, materiale TEXT,
            qty INTEGER, tipo TEXT, data TIMESTAMP DEFAULT NOW()
        )
    """)
    cur.execute("""
        CREATE TABLE IF NOT EXISTS calendario_eventi (
            id SERIAL PRIMARY KEY,
            data_evento DATE NOT NULL,
            titolo TEXT NOT NULL,
            categoria TEXT DEFAULT 'Altro',
            ricorrenza TEXT DEFAULT 'nessuna',
            ora_inizio TEXT DEFAULT '08:00',
            note TEXT DEFAULT '',
            created_at TIMESTAMP DEFAULT NOW()
        )
    """)
    cur.execute("""
        CREATE TABLE IF NOT EXISTS calendario_eccezioni (
            id SERIAL PRIMARY KEY,
            evento_id INTEGER REFERENCES calendario_eventi(id) ON DELETE CASCADE,
            data_eccezione DATE NOT NULL,
            UNIQUE(evento_id, data_eccezione)
        )
    """)
    cur.execute("""
        CREATE TABLE IF NOT EXISTS calendario_note_weekend (
            settimana_inizio DATE PRIMARY KEY,
            testo TEXT DEFAULT ''
        )
    """)
    cur.execute("""
        CREATE TABLE IF NOT EXISTS sessioni_cronometro (
            id SERIAL PRIMARY KEY,
            data_inizio TIMESTAMP NOT NULL,
            data_fine TIMESTAMP,
            ordini_json TEXT DEFAULT '[]',
            cambi_formato INTEGER DEFAULT 0,
            cambi_etichetta INTEGER DEFAULT 0,
            tempo_stimato_min INTEGER DEFAULT 0,
            tempo_reale_min INTEGER,
            stato TEXT DEFAULT 'in_corso'
        )
    """)
    conn.commit()
    cur.close()
    conn.close()

init_db()

# ==================================================
# CONFIGURAZIONE CLIENTI
# ==================================================

clients = {
    "Roberto": [
        "Catarratto 2L", "Rosato 2L", "Merlot 2L", "Il Nero 2L",
        "Bianco E.N. 2L", "Rosato E.N. 2L", "Rosso E.N. 2L",
        "Catarratto 1L", "Rosato 1L", "Merlot 1L",
        "Bianco S.E. 1L", "Rosato S.E. 1L", "Rosso S.E. 1L",
    ],
    "Francesco": [
        "Catarratto 2L", "Chardonnay 2L", "Rosato 2L", "Merlot 2L", "Syrah 2L",
        "Catarratto 1L", "Syrah 1L",
    ],
    "Emanuele": [
        "Catarratto 2L", "Rosato 2L", "Il Nero 2L", "Merlot 2L",
        "Vino Rosso 2L", "Syrah 2L",
        "Bianco S.E. 2L", "Rosato S.E. 2L", "Rosso S.E. 2L",
        "Catarratto R.B. 2L", "Rosato R.B. 2L", "Il Nero R.B. 2L", "Vino Rosso R.B. 2L",
        "Catarratto 1L", "Rosato 1L", "Il Nero 1L",
        "Bianco S.E. 1L", "Rosato S.E. 1L", "Rosso S.E. 1L",
    ],
    "Mazzarrone": [
        "Divino Bianco 2L", "Divino Rosato 2L", "Divino Rosso 2L", "Divino Syrah 2L",
        "Divino Bianco 1L", "Divino Rosato 1L", "Divino Rosso 1L", "Divino Syrah 1L",
        "Pachinos Bianco 2L", "Pachinos Rosato 2L", "Pachinos Rosso 2L", "Pachinos Syrah 2L",
        "Pachinos Bianco 1L", "Pachinos Rosato 1L", "Pachinos Rosso 1L", "Pachinos Syrah 1L",
    ],
    "Sisa": [
        "Bianco 2L", "Rosso 2L",
    ],
}

MOLTIPLICATORI = {
    "Roberto": {
        "Catarratto 2L": 6,  "Rosato 2L": 6,      "Merlot 2L": 6,
        "Il Nero 2L": 6,     "Bianco E.N. 2L": 6,  "Rosato E.N. 2L": 6,
        "Rosso E.N. 2L": 6,  "Catarratto 1L": 12,  "Rosato 1L": 12,
        "Merlot 1L": 12,     "Bianco S.E. 1L": 16, "Rosato S.E. 1L": 16,
        "Rosso S.E. 1L": 16,
    },
    "Francesco": {
        "Catarratto 2L": 6, "Chardonnay 2L": 6, "Rosato 2L": 6,
        "Merlot 2L": 6,     "Syrah 2L": 6,
        "Catarratto 1L": 12, "Syrah 1L": 12,
    },
    "Emanuele": {
        "Catarratto 2L": 9,      "Rosato 2L": 9,      "Il Nero 2L": 9,
        "Merlot 2L": 9,          "Vino Rosso 2L": 9,  "Syrah 2L": 9,
        "Bianco S.E. 2L": 9,     "Rosato S.E. 2L": 9, "Rosso S.E. 2L": 9,
        "Catarratto R.B. 2L": 9, "Rosato R.B. 2L": 9, "Il Nero R.B. 2L": 9,
        "Vino Rosso R.B. 2L": 9, "Catarratto 1L": 16, "Rosato 1L": 16,
        "Il Nero 1L": 16,        "Bianco S.E. 1L": 16,"Rosato S.E. 1L": 16,
        "Rosso S.E. 1L": 16,
    },
    "Mazzarrone": {
        "Divino Bianco 2L": 6,   "Divino Rosato 2L": 6,
        "Divino Rosso 2L": 6,    "Divino Syrah 2L": 6,
        "Divino Bianco 1L": 12,  "Divino Rosato 1L": 12,
        "Divino Rosso 1L": 12,   "Divino Syrah 1L": 12,
        "Pachinos Bianco 2L": 6, "Pachinos Rosato 2L": 6,
        "Pachinos Rosso 2L": 6,  "Pachinos Syrah 2L": 6,
        "Pachinos Bianco 1L": 12,"Pachinos Rosato 1L": 12,
        "Pachinos Rosso 1L": 12, "Pachinos Syrah 1L": 12,
    },
    "Sisa": {
        "Bianco 2L": 6, "Rosso 2L": 6,
    },
}

# Materie prime per cliente
# Bottiglie vuote: globali (non per cliente)
BOTTIGLIE_GLOBALI = ["Bottiglie 2L vuote", "Bottiglie 1L vuote"]

# Soglie minime fisse
SOGLIE_MP = {
    "Bottiglie 2L vuote": 8000,
    "Bottiglie 1L vuote": 1800,
}
SOGLIA_ETICHETTE = 1000

# Etichette per cliente
def _build_etichette(cliente):
    return [f"Etichetta {p}" for p in clients[cliente]]

ETICHETTE_CLIENTI = {c: _build_etichette(c) for c in clients}

# Fardelli per pedana per formato
PEDANE_CONFIG = {
    "Roberto":    {"2L": 64, "1L": 60},
    "Francesco":  {"2L": 64, "1L": 60},
    "Emanuele":   {"2L": 48, "1L": 48},
    "Mazzarrone": {"2L": 80, "1L": 60},
    "Sisa":       {"2L": 64, "1L": None},
}

# Tempi produzione per pedana (minuti) per formato
TEMPI_PEDANA = {
    "2L_64": 30,  # Roberto, Francesco, Sisa
    "2L_48": 35,  # Emanuele
    "2L_80": 45,  # Mazzarrone
    "1L_60": 30,  # Roberto, Francesco, Mazzarrone
    "1L_48": 35,  # Emanuele
}

def get_tempo_pedana(cliente, formato):
    """Ritorna minuti per pedana dato cliente e formato (2L o 1L)"""
    fard_ped = PEDANE_CONFIG.get(cliente, {}).get(formato)
    if not fard_ped:
        return 30
    key = f"{formato}_{fard_ped}"
    return TEMPI_PEDANA.get(key, 30)

def calcola_pedane(cliente, prodotto, qty_fardelli):
    """Calcola pedane da fardelli"""
    formato = "2L" if "2L" in prodotto else "1L"
    fard_ped = PEDANE_CONFIG.get(cliente, {}).get(formato, 64)
    if not fard_ped:
        return 0
    return qty_fardelli / fard_ped

CLIENTI_CONFIG = {
    "Roberto": {
        "bolla": "bolla_roberto.xlsx", "conteggio": "conteggio_roberto.xlsx",
        "prodotti": {
            "Catarratto 2L": 23,  "Rosato 2L": 25,      "Merlot 2L": 27,
            "Il Nero 2L": 29,     "Bianco E.N. 2L": 31,  "Rosato E.N. 2L": 33,
            "Rosso E.N. 2L": 35,  "Catarratto 1L": 37,   "Rosato 1L": 39,
            "Merlot 1L": 41,      "Bianco S.E. 1L": 43,  "Rosato S.E. 1L": 45,
            "Rosso S.E. 1L": 47,
        },
        "righe_az": [23,25,27,29,31,33,35,37,39,41,43,45,47],
        "cella_titolo_bolla": "H2",     "cella_data_bolla": "F55",
        "cella_titolo_conteggio": "G2", "cella_data_conteggio": "F53",
    },
    "Francesco": {
        "bolla": "bolla_francesco.xlsx", "conteggio": "conteggio_francesco.xlsx",
        "prodotti": {
            "Catarratto 2L": 21, "Chardonnay 2L": 23, "Rosato 2L": 25,
            "Merlot 2L": 27,     "Syrah 2L": 29,
            "Catarratto 1L": 43, "Syrah 1L": 45,
        },
        "righe_az": [21,23,25,27,29,35,37,39,43,45],
        "cella_titolo_bolla": "H2",     "cella_data_bolla": "F55",
        "cella_titolo_conteggio": "H2", "cella_data_conteggio": "F65",
    },
    "Emanuele": {
        "bolla": "bolla_emanuele.xlsx", "conteggio": "conteggio_emanuele.xlsx",
        "prodotti_bolla": {
            "Catarratto 2L": 21,      "Rosato 2L": 23,      "Il Nero 2L": 25,
            "Vino Rosso 2L": 27,      "Syrah 2L": 29,       "Merlot 2L": 31,
            "Catarratto 1L": 33,      "Rosato 1L": 35,      "Il Nero 1L": 37,
            "Catarratto R.B. 2L": 41, "Rosato R.B. 2L": 43, "Il Nero R.B. 2L": 45,
            "Vino Rosso R.B. 2L": 47, "Bianco S.E. 2L": 51, "Rosato S.E. 2L": 53,
            "Rosso S.E. 2L": 55,      "Bianco S.E. 1L": 57, "Rosato S.E. 1L": 59,
            "Rosso S.E. 1L": 61,
        },
        "prodotti_conteggio": {
            "Catarratto 2L": 25,      "Rosato 2L": 27,      "Il Nero 2L": 29,
            "Vino Rosso 2L": 31,      "Syrah 2L": 33,       "Merlot 2L": 35,
            "Catarratto 1L": 37,      "Rosato 1L": 39,      "Il Nero 1L": 41,
            "Catarratto R.B. 2L": 45, "Rosato R.B. 2L": 47, "Il Nero R.B. 2L": 49,
            "Vino Rosso R.B. 2L": 51, "Bianco S.E. 2L": 55, "Rosato S.E. 2L": 57,
            "Rosso S.E. 2L": 59,      "Bianco S.E. 1L": 61, "Rosato S.E. 1L": 63,
            "Rosso S.E. 1L": 65,
        },
        "righe_az_bolla":     [21,23,25,27,29,31,33,35,37,41,43,45,47,51,53,55,57,59,61],
        "righe_az_conteggio": [25,27,29,31,33,35,37,39,41,45,47,49,51,55,57,59,61,63,65],
        "cella_titolo_bolla": "H2",     "cella_data_bolla": "F67",
        "cella_titolo_conteggio": "H2", "cella_data_conteggio": "F71",
    },
    "Mazzarrone": {
        "bolla": None, "conteggio": None, "prodotti": {}, "righe_az": [],
        "cella_titolo_bolla": None, "cella_data_bolla": None,
        "cella_titolo_conteggio": None, "cella_data_conteggio": None,
    },
    "Sisa": {
        "bolla": None, "conteggio": None, "prodotti": {}, "righe_az": [],
        "cella_titolo_bolla": None, "cella_data_bolla": None,
        "cella_titolo_conteggio": None, "cella_data_conteggio": None,
    },
}


# ==================================================
# FUNZIONI EXCEL
# ==================================================

def _aggiorna_excel(ws, cliente, richieste_fardelli, tipo="bolla"):
    cfg = CLIENTI_CONFIG[cliente]
    if cliente == "Emanuele":
        mappa    = cfg["prodotti_bolla"] if tipo == "bolla" else cfg["prodotti_conteggio"]
        righe_az = cfg["righe_az_bolla"] if tipo == "bolla" else cfg["righe_az_conteggio"]
    else:
        mappa    = cfg.get("prodotti", {})
        righe_az = cfg.get("righe_az", [])
    for riga in righe_az:
        ws[f"G{riga}"] = 0
    for prodotto, fardelli in richieste_fardelli:
        if prodotto in mappa:
            ws[f"G{mappa[prodotto]}"] = fardelli


def _genera_file(cliente, richieste_fardelli, tipo):
    cfg = CLIENTI_CONFIG[cliente]
    nome_modello = cfg["bolla"] if tipo == "bolla" else cfg["conteggio"]
    if not nome_modello:
        return None, "Modello non ancora disponibile per questo cliente"
    file_modello = os.path.join(MODELLI_DIR, nome_modello)
    if not os.path.exists(file_modello):
        return None, f"File modello mancante: {nome_modello}"
    wb = load_workbook(file_modello)
    ws = wb.active
    cella_titolo = cfg[f"cella_titolo_{tipo}"]
    cella_data   = cfg[f"cella_data_{tipo}"]
    if cella_titolo:
        ws[cella_titolo] = "DOCUMENTO DI TRASPORTO\nN.          DEL\n"
    if cella_data:
        ws[cella_data] = "DATA RITIRO\n\n\n"
    _aggiorna_excel(ws, cliente, richieste_fardelli, tipo)
    output = os.path.join(BASE_DIR, f"{tipo}_generato_{cliente.lower()}.xlsx")
    wb.save(output)
    return output, None


def _leggi_richieste_fardelli(cliente, form):
    richieste = []
    for i, prodotto in enumerate(clients[cliente]):
        val = form.get(f"qty_{i}")
        if val and val.isdigit():
            f = int(val)
            if f > 0:
                richieste.append((prodotto, f))
    return richieste


def _fardelli_a_bottiglie(cliente, richieste_fardelli):
    molt = MOLTIPLICATORI.get(cliente, {})
    return [(p, f * molt.get(p, 1)) for p, f in richieste_fardelli]


def _is_2L(prodotto):
    return "2L" in prodotto or "2l" in prodotto

def _is_1L(prodotto):
    return "1L" in prodotto or "1l" in prodotto


def _scarico_automatico_bottiglie(cur, cliente, prodotti_bottiglie):
    """
    Scarica automaticamente:
    - Bottiglie vuote: globali (cliente="GLOBALE"), aggregate per formato
    - Etichette: per cliente, una per bottiglia per ogni prodotto
    prodotti_bottiglie = [(prodotto, qty_bt), ...]
    """
    # --- Bottiglie vuote GLOBALI ---
    bt2L = sum(q for p, q in prodotti_bottiglie if _is_2L(p))
    bt1L = sum(q for p, q in prodotti_bottiglie if _is_1L(p))

    for materiale, qty_usate in [("Bottiglie 2L vuote", bt2L), ("Bottiglie 1L vuote", bt1L)]:
        if qty_usate <= 0:
            continue
        cur.execute(
            "SELECT * FROM materie_prime WHERE cliente=%s AND materiale=%s",
            ("GLOBALE", materiale)
        )
        row = cur.fetchone()
        if row and row["qty"] >= qty_usate:
            cur.execute("UPDATE materie_prime SET qty=%s WHERE id=%s",
                        (row["qty"] - qty_usate, row["id"]))
            cur.execute(
                "INSERT INTO storico_mp(cliente, materiale, qty, tipo) VALUES(%s,%s,%s,%s)",
                ("GLOBALE", materiale, qty_usate, "Scarico automatico produzione")
            )

    # --- Etichette per CLIENTE ---
    for prodotto, qty_bt in prodotti_bottiglie:
        materiale = f"Etichetta {prodotto}"
        cur.execute(
            "SELECT * FROM materie_prime WHERE cliente=%s AND materiale=%s",
            (cliente, materiale)
        )
        row = cur.fetchone()
        if row and row["qty"] >= qty_bt:
            cur.execute("UPDATE materie_prime SET qty=%s WHERE id=%s",
                        (row["qty"] - qty_bt, row["id"]))
            cur.execute(
                "INSERT INTO storico_mp(cliente, materiale, qty, tipo) VALUES(%s,%s,%s,%s)",
                (cliente, materiale, qty_bt, "Scarico automatico produzione")
            )


def _get_alert_mp():
    """Ritorna lista di dict con tutti gli alert sotto soglia (soglie fisse)."""
    try:
        conn = db()
        cur = conn.cursor()
        cur.execute("SELECT * FROM materie_prime ORDER BY cliente, materiale")
        rows = cur.fetchall()
        cur.close()
        conn.close()
        alert = []
        for r in rows:
            soglia = SOGLIE_MP.get(r["materiale"], SOGLIA_ETICHETTE)
            if r["qty"] <= soglia:
                alert.append({
                    "cliente":   r["cliente"],
                    "materiale": r["materiale"],
                    "qty":       r["qty"],
                    "soglia":    soglia,
                })
        return alert
    except:
        return []


def _conta_alert_mp():
    return len(_get_alert_mp())


# ==================================================
# HOME
# ==================================================
@app.route("/")
def home():
    from datetime import timedelta
    oggi = now_rome().date()
    lun = oggi - __import__('datetime').timedelta(days=oggi.weekday())

    conn = db()
    cur = conn.cursor()
    cur.execute("""
        SELECT * FROM calendario_eventi
        WHERE (ricorrenza='nessuna' AND data_evento>=%s AND data_evento<=%s)
           OR ricorrenza IN ('settimanale','bisettimanale')
        ORDER BY data_evento, ora_inizio
    """, (lun, lun + timedelta(days=6)))
    tutti = cur.fetchall()
    cur.close()
    conn.close()

    # Costruisce dict giorni settimana
    giorni_sett = [lun + timedelta(days=i) for i in range(7)]
    eventi_home = {}
    for g in giorni_sett:
        eventi_home[g.isoformat()] = []

    conn2 = db()
    cur2 = conn2.cursor()
    for e in tutti:
        if e["ricorrenza"] == "nessuna":
            k = e["data_evento"].isoformat()
            if k in eventi_home:
                eventi_home[k].append(dict(e))
        else:
            for g in giorni_sett:
                if g.weekday() == e["data_evento"].weekday() and g >= e["data_evento"]:
                    if e["ricorrenza"] == "bisettimanale":
                        if ((g - e["data_evento"]).days // 7) % 2 != 0:
                            continue
                    cur2.execute(
                        "SELECT id FROM calendario_eccezioni WHERE evento_id=%s AND data_eccezione=%s",
                        (e["id"], g)
                    )
                    if not cur2.fetchone():
                        ev = dict(e); ev["data_evento"] = g
                        eventi_home[g.isoformat()].append(ev)
    cur2.close()
    conn2.close()

    return render_template("home.html",
        alert_mp=_get_alert_mp(),
        giorni_sett=giorni_sett,
        eventi_home=eventi_home,
        oggi=oggi,
    )


# ==================================================
# STORICO
# ==================================================
@app.route("/storico")
def storico():
    from datetime import datetime, timedelta

    periodo = request.args.get("periodo", "30")
    data_da = request.args.get("data_da", "")
    data_a  = request.args.get("data_a", "")
    oggi    = datetime.now().date()

    if periodo == "7":
        filtro_da = oggi - timedelta(days=7)
        filtro_a  = oggi + timedelta(days=1)
    elif periodo == "30":
        filtro_da = oggi - timedelta(days=30)
        filtro_a  = oggi + timedelta(days=1)
    elif periodo == "custom" and data_da and data_a:
        try:
            filtro_da = datetime.strptime(data_da, "%Y-%m-%d").date()
            filtro_a  = datetime.strptime(data_a, "%Y-%m-%d").date() + timedelta(days=1)
        except:
            filtro_da = oggi - timedelta(days=30)
            filtro_a  = oggi + timedelta(days=1)
    else:
        filtro_da = None
        filtro_a  = None

    conn = db()
    cur = conn.cursor()

    def fetch_per_cliente(tipo):
        if filtro_da:
            cur.execute(
                "SELECT * FROM storico WHERE tipo=%s AND data>=%s AND data<%s ORDER BY data DESC",
                (tipo, filtro_da, filtro_a)
            )
        else:
            cur.execute("SELECT * FROM storico WHERE tipo=%s ORDER BY data DESC", (tipo,))
        grouped = {}
        for r in cur.fetchall():
            grouped.setdefault(r["cliente"], []).append(r)
        return grouped

    prod_per_cliente = fetch_per_cliente("Produzione Inserita")
    mag_per_cliente  = fetch_per_cliente("Passato a Magazzino")
    scar_per_cliente = fetch_per_cliente("Scarico Magazzino")
    cons_per_cliente = fetch_per_cliente("Consegna")

    if filtro_da:
        cur.execute(
            "SELECT * FROM storico_mp WHERE data>=%s AND data<%s ORDER BY data DESC",
            (filtro_da, filtro_a)
        )
    else:
        cur.execute("SELECT * FROM storico_mp ORDER BY data DESC")
    rows_mp = cur.fetchall()

    cur.close()
    conn.close()

    return render_template("storico.html",
        prod_per_cliente=prod_per_cliente,
        mag_per_cliente=mag_per_cliente,
        scar_per_cliente=scar_per_cliente,
        cons_per_cliente=cons_per_cliente,
        rows_mp=rows_mp,
        periodo=periodo,
        data_da=data_da,
        data_a=data_a,
    )


# ==================================================
# PRODUZIONE
# ==================================================
@app.route("/produzione")
def produzione():
    conn = db()
    cur = conn.cursor()
    cur.execute("SELECT * FROM produzione ORDER BY id")
    righe = cur.fetchall()
    cur.execute("SELECT * FROM note ORDER BY data DESC")
    note = cur.fetchall()
    # Sessione cronometro attiva
    cur.execute("""
        SELECT * FROM sessioni_cronometro
        WHERE stato='in_corso' ORDER BY data_inizio DESC LIMIT 1
    """)
    sessione_attiva = cur.fetchone()
    cur.close()
    conn.close()

    # Calcola stima - solo ordini NON ancora completati (done=0)
    stima_min = 0
    dettaglio_stima = []
    prodotti_unici = set()
    formati = set()

    for r in righe:
        if r["done"] == 1:
            continue
        molt = MOLTIPLICATORI.get(r["cliente"], {}).get(r["prodotto"], 1)
        fardelli = r["qty"] / molt
        formato = "2L" if "2L" in r["prodotto"] else "1L"
        fard_ped = PEDANE_CONFIG.get(r["cliente"], {}).get(formato, 64)
        mins = 0
        if fard_ped:
            pedane = fardelli / fard_ped
            mins = round(pedane * get_tempo_pedana(r["cliente"], formato))
            stima_min += mins
        prodotti_unici.add(r["prodotto"])
        formati.add(formato)
        dettaglio_stima.append({
            "cliente": r["cliente"],
            "prodotto": r["prodotto"],
            "fardelli": round(fardelli, 1),
            "mins": mins,
        })

    fissi_min = 90  # riscaldamento 30 + pranzo 60
    cambi_et_min = len(prodotti_unici) * 3
    cambio_fmt_min = 10 if len(formati) > 1 else 0
    base = stima_min + fissi_min + cambi_et_min + cambio_fmt_min
    imprevisti_min = round(base * 0.075)
    stima_min = round(base * 1.075)

    return render_template("produzione.html", clients=clients, rows=righe,
                           note=note, moltiplicatori=MOLTIPLICATORI,
                           sessione_attiva=sessione_attiva,
                           stima_min=stima_min,
                           dettaglio_stima=dettaglio_stima,
                           fissi_min=fissi_min,
                           cambi_et_min=cambi_et_min,
                           cambio_fmt_min=cambio_fmt_min,
                           imprevisti_min=imprevisti_min)


@app.route("/nuova_produzione", methods=["POST"])
def nuova_produzione():
    cliente = request.form["client"]
    conn = db()
    cur = conn.cursor()
    molt = MOLTIPLICATORI.get(cliente, {})
    for i, prodotto in enumerate(clients[cliente]):
        val = request.form.get(f"qty_{i}")
        if val and val.isdigit():
            f = int(val)
            if f > 0:
                bottiglie = f * molt.get(prodotto, 1)
                cur.execute(
                    "INSERT INTO produzione(cliente, prodotto, qty) VALUES(%s,%s,%s)",
                    (cliente, prodotto, bottiglie)
                )
                cur.execute(
                    "INSERT INTO storico(cliente, prodotto, qty, tipo) VALUES(%s,%s,%s,%s)",
                    (cliente, prodotto, bottiglie, "Produzione Inserita")
                )
    conn.commit()
    cur.close()
    conn.close()
    return redirect("/produzione")


@app.route("/toggle/<int:id>")
def toggle(id):
    conn = db()
    cur = conn.cursor()
    cur.execute("SELECT done FROM produzione WHERE id=%s", (id,))
    row = cur.fetchone()
    nuovo = 0 if row["done"] == 1 else 1
    if nuovo == 1:
        cur.execute(
            "UPDATE produzione SET done=%s, timestamp_done=%s WHERE id=%s",
            (nuovo, now_rome(), id)
        )
    else:
        cur.execute(
            "UPDATE produzione SET done=%s, timestamp_done=NULL WHERE id=%s",
            (nuovo, id)
        )
    conn.commit()
    cur.close()
    conn.close()
    return redirect("/produzione")


@app.route("/passa_magazzino")
def passa_magazzino():
    conn = db()
    cur = conn.cursor()
    cur.execute("SELECT * FROM produzione WHERE done=1")
    finiti = cur.fetchall()

    # Raggruppa per cliente per lo scarico bottiglie
    per_cliente = {}
    for r in finiti:
        per_cliente.setdefault(r["cliente"], []).append((r["prodotto"], r["qty"]))

    for r in finiti:
        cur.execute(
            "SELECT * FROM stock WHERE cliente=%s AND prodotto=%s",
            (r["cliente"], r["prodotto"])
        )
        ex = cur.fetchone()
        if ex:
            cur.execute("UPDATE stock SET qty=%s WHERE id=%s",
                        (ex["qty"] + r["qty"], ex["id"]))
        else:
            cur.execute("INSERT INTO stock(cliente, prodotto, qty) VALUES(%s,%s,%s)",
                        (r["cliente"], r["prodotto"], r["qty"]))
        cur.execute(
            "INSERT INTO storico(cliente, prodotto, qty, tipo) VALUES(%s,%s,%s,%s)",
            (r["cliente"], r["prodotto"], r["qty"], "Passato a Magazzino")
        )
        cur.execute("DELETE FROM produzione WHERE id=%s", (r["id"],))

    # Scarico automatico bottiglie vuote per cliente
    for cliente, prodotti_bt in per_cliente.items():
        _scarico_automatico_bottiglie(cur, cliente, prodotti_bt)

    conn.commit()
    cur.close()
    conn.close()
    return redirect("/produzione")


@app.route("/aggiungi_nota", methods=["POST"])
def aggiungi_nota():
    testo = request.form.get("testo", "").strip()
    if testo:
        conn = db()
        cur = conn.cursor()
        cur.execute("INSERT INTO note(testo) VALUES(%s)", (testo,))
        conn.commit()
        cur.close()
        conn.close()
    return redirect("/produzione")


@app.route("/elimina_nota/<int:id>")
def elimina_nota(id):
    conn = db()
    cur = conn.cursor()
    cur.execute("DELETE FROM note WHERE id=%s", (id,))
    conn.commit()
    cur.close()
    conn.close()
    return redirect("/produzione")


# ==================================================
# MAGAZZINO PRODOTTI FINITI
# ==================================================
@app.route("/magazzino")
def magazzino():
    msg = request.args.get("msg", "")
    cliente_sel = request.args.get("cliente", "")
    conn = db()
    cur = conn.cursor()
    cur.execute("SELECT * FROM stock WHERE qty > 0 ORDER BY cliente, prodotto")
    stock_rows = cur.fetchall()
    cur.execute("SELECT * FROM materie_prime ORDER BY cliente, materiale")
    mp_rows = cur.fetchall()
    cur.close()
    conn.close()

    # Stock prodotti finiti per cliente
    grouped = {}
    for r in stock_rows:
        grouped.setdefault(r["cliente"], []).append(r)

    # Materie prime per cliente + alert
    grouped_mp = {}
    alert_mp = []
    for r in mp_rows:
        grouped_mp.setdefault(r["cliente"], []).append(r)
        if r["soglia_minima"] > 0 and r["qty"] <= r["soglia_minima"]:
            alert_mp.append(r)

    # Separa giacenze bottiglie globali da etichette per cliente
    bottiglie_globali = grouped_mp.get("GLOBALE", [])
    grouped_etichette = {k: v for k, v in grouped_mp.items() if k != "GLOBALE"}

    return render_template(
        "magazzino.html",
        grouped=grouped,
        bottiglie_globali=bottiglie_globali,
        grouped_etichette=grouped_etichette,
        alert_mp=_get_alert_mp(),
        clients=clients,
        etichette_clienti=ETICHETTE_CLIENTI,
        msg=msg,
        cliente_sel=cliente_sel,
        moltiplicatori=MOLTIPLICATORI,
        bt_pedana_2l=BT_PER_PEDANA_2L,
        bt_pedana_1l=BT_PER_PEDANA_1L,
    )


@app.route("/scarica", methods=["POST"])
def scarica():
    cliente = request.form["client"]
    molt = MOLTIPLICATORI.get(cliente, {})
    richieste_bt = []
    for i, prodotto in enumerate(clients[cliente]):
        val = request.form.get(f"qty_{i}")
        if val and val.isdigit():
            f = int(val)
            if f > 0:
                richieste_bt.append((prodotto, f * molt.get(prodotto, 1)))
    if not richieste_bt:
        return redirect("/magazzino?msg=Nessun prodotto selezionato&cliente=" + cliente)
    conn = db()
    cur = conn.cursor()
    for prodotto, q in richieste_bt:
        cur.execute("SELECT * FROM stock WHERE cliente=%s AND prodotto=%s", (cliente, prodotto))
        row = cur.fetchone()
        if not row:
            cur.close(); conn.close()
            return redirect("/magazzino?msg=" + prodotto + " non presente&cliente=" + cliente)
        if row["qty"] < q:
            cur.close(); conn.close()
            return redirect("/magazzino?msg=" + prodotto + " quantita insufficiente&cliente=" + cliente)
    for prodotto, q in richieste_bt:
        cur.execute("SELECT * FROM stock WHERE cliente=%s AND prodotto=%s", (cliente, prodotto))
        row = cur.fetchone()
        nuova = row["qty"] - q
        if nuova == 0:
            cur.execute("DELETE FROM stock WHERE id=%s", (row["id"],))
        else:
            cur.execute("UPDATE stock SET qty=%s WHERE id=%s", (nuova, row["id"]))
        cur.execute(
            "INSERT INTO storico(cliente, prodotto, qty, tipo) VALUES(%s,%s,%s,%s)",
            (cliente, prodotto, q, "Scarico Magazzino")
        )
    conn.commit()
    cur.close()
    conn.close()
    return redirect("/magazzino?msg=Scarico completato&cliente=" + cliente)


# Materie prime: inizializza, carico, scarico, soglia
@app.route("/init_materie_prime")
def init_materie_prime():
    conn = db()
    cur = conn.cursor()
    # Bottiglie vuote: globali
    for materiale in BOTTIGLIE_GLOBALI:
        cur.execute(
            "SELECT id FROM materie_prime WHERE cliente=%s AND materiale=%s",
            ("GLOBALE", materiale)
        )
        if not cur.fetchone():
            cur.execute(
                "INSERT INTO materie_prime(cliente, materiale, qty, soglia_minima) VALUES(%s,%s,0,0)",
                ("GLOBALE", materiale)
            )
    # Etichette: per cliente
    for cliente, etichette in ETICHETTE_CLIENTI.items():
        for materiale in etichette:
            cur.execute(
                "SELECT id FROM materie_prime WHERE cliente=%s AND materiale=%s",
                (cliente, materiale)
            )
            if not cur.fetchone():
                cur.execute(
                    "INSERT INTO materie_prime(cliente, materiale, qty, soglia_minima) VALUES(%s,%s,0,0)",
                    (cliente, materiale)
                )
    conn.commit()
    cur.close()
    conn.close()
    return redirect("/magazzino?msg=Materie prime inizializzate")


@app.route("/carico_mp", methods=["POST"])
def carico_mp():
    cliente = request.form.get("cliente", "GLOBALE")
    conn = db()
    cur = conn.cursor()
    caricati = 0

    # Bottiglie globali (da form dedicato con unita pedane/pezzi)
    for mat in BOTTIGLIE_GLOBALI:
        campo = mat.replace(" ", "_")
        val   = request.form.get(f"qty_{campo}", "")
        unita = request.form.get(f"unita_{campo}", "pezzi")
        if val and val.isdigit() and int(val) > 0:
            q = int(val)
            if unita == "pedane":
                q = q * (BT_PER_PEDANA_2L if "2L" in mat else BT_PER_PEDANA_1L)
            cur.execute(
                "SELECT * FROM materie_prime WHERE cliente=%s AND materiale=%s",
                ("GLOBALE", mat)
            )
            row = cur.fetchone()
            if row:
                cur.execute("UPDATE materie_prime SET qty=%s WHERE id=%s",
                            (row["qty"] + q, row["id"]))
            else:
                cur.execute(
                    "INSERT INTO materie_prime(cliente, materiale, qty, soglia_minima) VALUES(%s,%s,%s,0)",
                    ("GLOBALE", mat, q)
                )
            cur.execute(
                "INSERT INTO storico_mp(cliente, materiale, qty, tipo) VALUES(%s,%s,%s,%s)",
                ("GLOBALE", mat, q, "Carico")
            )
            caricati += 1

    # Etichette per cliente (tutte insieme)
    if cliente != "GLOBALE" and cliente in ETICHETTE_CLIENTI:
        for etichetta in ETICHETTE_CLIENTI[cliente]:
            campo = str(ETICHETTE_CLIENTI[cliente].index(etichetta))
            val   = request.form.get(f"qty_et_{campo}", "")
            if val and val.isdigit() and int(val) > 0:
                q = int(val)
                cur.execute(
                    "SELECT * FROM materie_prime WHERE cliente=%s AND materiale=%s",
                    (cliente, etichetta)
                )
                row = cur.fetchone()
                if row:
                    cur.execute("UPDATE materie_prime SET qty=%s WHERE id=%s",
                                (row["qty"] + q, row["id"]))
                else:
                    cur.execute(
                        "INSERT INTO materie_prime(cliente, materiale, qty, soglia_minima) VALUES(%s,%s,%s,0)",
                        (cliente, etichetta, q)
                    )
                cur.execute(
                    "INSERT INTO storico_mp(cliente, materiale, qty, tipo) VALUES(%s,%s,%s,%s)",
                    (cliente, etichetta, q, "Carico")
                )
                caricati += 1

    conn.commit()
    cur.close()
    conn.close()
    msg = "Carico registrato" if caricati > 0 else "Nessuna quantità inserita"
    return redirect("/magazzino?msg=" + msg + "&cliente=" + cliente)


@app.route("/scarico_mp", methods=["POST"])
def scarico_mp():
    materiale = request.form["materiale"]
    cliente = "GLOBALE" if materiale in BOTTIGLIE_GLOBALI else request.form["cliente"]
    val       = request.form.get("qty", "0")
    if not val.isdigit() or int(val) <= 0:
        return redirect("/magazzino?msg=Quantita non valida&cliente=" + cliente)
    q = int(val)
    conn = db()
    cur = conn.cursor()
    cur.execute(
        "SELECT * FROM materie_prime WHERE cliente=%s AND materiale=%s",
        (cliente, materiale)
    )
    row = cur.fetchone()
    if not row or row["qty"] < q:
        cur.close(); conn.close()
        return redirect("/magazzino?msg=" + materiale + " quantita insufficiente&cliente=" + cliente)
    cur.execute("UPDATE materie_prime SET qty=%s WHERE id=%s", (row["qty"] - q, row["id"]))
    cur.execute(
        "INSERT INTO storico_mp(cliente, materiale, qty, tipo) VALUES(%s,%s,%s,%s)",
        (cliente, materiale, q, "Scarico")
    )
    conn.commit()
    cur.close()
    conn.close()
    return redirect("/magazzino?msg=Scarico registrato&cliente=" + cliente)


@app.route("/set_soglia_mp", methods=["POST"])
def set_soglia_mp():
    cliente   = request.form["cliente"]
    materiale = request.form["materiale"]
    val       = request.form.get("soglia", "0")
    if not val.isdigit():
        return redirect("/magazzino?msg=Soglia non valida&cliente=" + cliente)
    conn = db()
    cur = conn.cursor()
    cur.execute(
        "UPDATE materie_prime SET soglia_minima=%s WHERE cliente=%s AND materiale=%s",
        (int(val), cliente, materiale)
    )
    conn.commit()
    cur.close()
    conn.close()
    return redirect("/magazzino?msg=Soglia aggiornata&cliente=" + cliente)


# ==================================================
# CONSEGNE
# ==================================================
@app.route("/consegne")
def consegne():
    msg = request.args.get("msg", "")
    cliente_sel = request.args.get("cliente", "")
    conn = db()
    cur = conn.cursor()
    cur.execute("SELECT * FROM stock WHERE qty > 0 ORDER BY cliente, prodotto")
    rows = cur.fetchall()
    cur.close()
    conn.close()
    grouped = {}
    for r in rows:
        grouped.setdefault(r["cliente"], []).append(r)
    return render_template("consegne.html", clients=clients, grouped=grouped,
                           cliente_sel=cliente_sel, msg=msg,
                           moltiplicatori=MOLTIPLICATORI)


@app.route("/esegui_consegna", methods=["POST"])
def esegui_consegna():
    cliente = request.form["client"]
    richieste_f = _leggi_richieste_fardelli(cliente, request.form)
    if not richieste_f:
        return redirect("/consegne?msg=Nessun prodotto selezionato&cliente=" + cliente)
    richieste_bt = _fardelli_a_bottiglie(cliente, richieste_f)
    conn = db()
    cur = conn.cursor()
    for prodotto, q in richieste_bt:
        cur.execute("SELECT * FROM stock WHERE cliente=%s AND prodotto=%s", (cliente, prodotto))
        row = cur.fetchone()
        if not row:
            cur.close(); conn.close()
            return redirect("/consegne?msg=" + prodotto + " non presente&cliente=" + cliente)
        if row["qty"] < q:
            cur.close(); conn.close()
            return redirect("/consegne?msg=" + prodotto + " quantita insufficiente&cliente=" + cliente)
    for prodotto, q in richieste_bt:
        cur.execute("SELECT * FROM stock WHERE cliente=%s AND prodotto=%s", (cliente, prodotto))
        row = cur.fetchone()
        nuova = row["qty"] - q
        if nuova == 0:
            cur.execute("DELETE FROM stock WHERE id=%s", (row["id"],))
        else:
            cur.execute("UPDATE stock SET qty=%s WHERE id=%s", (nuova, row["id"]))
        cur.execute(
            "INSERT INTO storico(cliente, prodotto, qty, tipo) VALUES(%s,%s,%s,%s)",
            (cliente, prodotto, q, "Consegna")
        )
    conn.commit()
    cur.close()
    conn.close()
    session["consegna_cliente"]   = cliente
    session["consegna_fardelli"]  = json.dumps(richieste_f)
    session["consegna_bottiglie"] = json.dumps(richieste_bt)
    return redirect("/conferma_consegna")


@app.route("/conferma_consegna")
def conferma_consegna():
    cliente      = session.get("consegna_cliente", "")
    richieste_f  = json.loads(session.get("consegna_fardelli", "[]"))
    richieste_bt = json.loads(session.get("consegna_bottiglie", "[]"))
    if not cliente or not richieste_f:
        return redirect("/consegne?msg=Nessuna consegna attiva")
    return render_template("conferma_consegna.html", cliente=cliente,
                           richieste_f=richieste_f, richieste_bt=richieste_bt)


@app.route("/download_bolla")
def download_bolla():
    cliente     = session.get("consegna_cliente", "")
    richieste_f = json.loads(session.get("consegna_fardelli", "[]"))
    if not cliente or not richieste_f:
        return redirect("/consegne?msg=Nessuna consegna attiva")
    output, errore = _genera_file(cliente, richieste_f, "bolla")
    if errore:
        return redirect("/conferma_consegna?msg=" + errore)
    return send_file(output, as_attachment=True, download_name=f"Bolla_{cliente}.xlsx")


@app.route("/download_conteggio")
def download_conteggio():
    cliente     = session.get("consegna_cliente", "")
    richieste_f = json.loads(session.get("consegna_fardelli", "[]"))
    if not cliente or not richieste_f:
        return redirect("/consegne?msg=Nessuna consegna attiva")
    output, errore = _genera_file(cliente, richieste_f, "conteggio")
    if errore:
        return redirect("/conferma_consegna?msg=" + errore)
    return send_file(output, as_attachment=True, download_name=f"Conteggio_{cliente}.xlsx")


@app.route("/solo_bolla", methods=["POST"])
def solo_bolla():
    cliente = request.form["client"]
    richieste_f = _leggi_richieste_fardelli(cliente, request.form)
    if not richieste_f:
        return redirect("/consegne?msg=Nessun prodotto selezionato&cliente=" + cliente)
    output, errore = _genera_file(cliente, richieste_f, "bolla")
    if errore:
        return redirect("/consegne?msg=" + errore + "&cliente=" + cliente)
    return send_file(output, as_attachment=True, download_name=f"Bolla_{cliente}.xlsx")


@app.route("/solo_conteggio", methods=["POST"])
def solo_conteggio():
    cliente = request.form["client"]
    richieste_f = _leggi_richieste_fardelli(cliente, request.form)
    if not richieste_f:
        return redirect("/consegne?msg=Nessun prodotto selezionato&cliente=" + cliente)
    output, errore = _genera_file(cliente, richieste_f, "conteggio")
    if errore:
        return redirect("/consegne?msg=" + errore + "&cliente=" + cliente)
    return send_file(output, as_attachment=True, download_name=f"Conteggio_{cliente}.xlsx")


# ==================================================
# BACKUP
# ==================================================
@app.route("/backup")
def backup():
    """Esporta tutto il database in JSON scaricabile."""
    from flask import jsonify
    import datetime

    conn = db()
    cur = conn.cursor()

    cur.execute("SELECT * FROM stock ORDER BY cliente, prodotto")
    stock = [dict(r) for r in cur.fetchall()]

    cur.execute("SELECT * FROM produzione ORDER BY id")
    produzione_rows = [dict(r) for r in cur.fetchall()]

    cur.execute("SELECT * FROM storico ORDER BY data DESC")
    storico_rows = [dict(r) for r in cur.fetchall()]

    cur.execute("SELECT * FROM materie_prime ORDER BY cliente, materiale")
    mp_rows = [dict(r) for r in cur.fetchall()]

    cur.execute("SELECT * FROM storico_mp ORDER BY data DESC")
    storico_mp_rows = [dict(r) for r in cur.fetchall()]

    cur.execute("SELECT * FROM note ORDER BY data DESC")
    note_rows = [dict(r) for r in cur.fetchall()]

    cur.close()
    conn.close()

    # Converti datetime in stringa per JSON
    def serialize(rows):
        for r in rows:
            for k, v in r.items():
                if hasattr(v, "isoformat"):
                    r[k] = v.isoformat()
        return rows

    data = {
        "backup_data": datetime.datetime.now().isoformat(),
        "stock": serialize(stock),
        "produzione": serialize(produzione_rows),
        "storico": serialize(storico_rows),
        "materie_prime": serialize(mp_rows),
        "storico_mp": serialize(storico_mp_rows),
        "note": serialize(note_rows),
    }

    response = jsonify(data)
    response.headers["Content-Disposition"] = (
        f"attachment; filename=backup_{datetime.date.today()}.json"
    )
    return response


# ==================================================
# ANALISI
# ==================================================
@app.route("/analisi")
def analisi():
    from datetime import datetime, timedelta

    periodo = request.args.get("periodo", "30")
    data_da = request.args.get("data_da", "")
    data_a  = request.args.get("data_a", "")
    oggi    = datetime.now().date()

    if periodo == "7":
        filtro_da = oggi - timedelta(days=7)
        filtro_a  = oggi + timedelta(days=1)
    elif periodo == "30":
        filtro_da = oggi - timedelta(days=30)
        filtro_a  = oggi + timedelta(days=1)
    elif periodo == "custom" and data_da and data_a:
        try:
            filtro_da = datetime.strptime(data_da, "%Y-%m-%d").date()
            filtro_a  = datetime.strptime(data_a, "%Y-%m-%d").date() + timedelta(days=1)
        except:
            filtro_da = oggi - timedelta(days=30)
            filtro_a  = oggi + timedelta(days=1)
    else:
        filtro_da = oggi - timedelta(days=30)
        filtro_a  = oggi + timedelta(days=1)

    conn = db()
    cur = conn.cursor()

    # 1. Produzione per cliente (bottiglie totali)
    cur.execute("""
        SELECT cliente, SUM(qty) as totale FROM storico
        WHERE tipo='Produzione Inserita' AND data>=%s AND data<%s
        GROUP BY cliente ORDER BY totale DESC
    """, (filtro_da, filtro_a))
    prod_cliente = cur.fetchall()

    # 2. Consegne per cliente (bottiglie totali)
    cur.execute("""
        SELECT cliente, SUM(qty) as totale FROM storico
        WHERE tipo='Consegna' AND data>=%s AND data<%s
        GROUP BY cliente ORDER BY totale DESC
    """, (filtro_da, filtro_a))
    cons_cliente = cur.fetchall()

    # 3. Produzione per prodotto (top 10)
    cur.execute("""
        SELECT prodotto, cliente, SUM(qty) as totale FROM storico
        WHERE tipo='Produzione Inserita' AND data>=%s AND data<%s
        GROUP BY prodotto, cliente ORDER BY totale DESC LIMIT 10
    """, (filtro_da, filtro_a))
    prod_prodotto = cur.fetchall()

    # 4. Trend produzione giornaliero
    cur.execute("""
        SELECT DATE(data) as giorno, SUM(qty) as totale FROM storico
        WHERE tipo='Produzione Inserita' AND data>=%s AND data<%s
        GROUP BY DATE(data) ORDER BY giorno
    """, (filtro_da, filtro_a))
    trend_prod = cur.fetchall()

    # 5. Trend consegne giornaliero
    cur.execute("""
        SELECT DATE(data) as giorno, SUM(qty) as totale FROM storico
        WHERE tipo='Consegna' AND data>=%s AND data<%s
        GROUP BY DATE(data) ORDER BY giorno
    """, (filtro_da, filtro_a))
    trend_cons = cur.fetchall()

    # 6. Consumo materie prime bottiglie + proiezione
    cur.execute("""
        SELECT materiale, SUM(qty) as totale FROM storico_mp
        WHERE tipo='Scarico automatico produzione' AND data>=%s AND data<%s
        GROUP BY materiale
    """, (filtro_da, filtro_a))
    consumo_bt = {r["materiale"]: r["totale"] for r in cur.fetchall()}

    cur.execute("""
        SELECT * FROM materie_prime WHERE cliente='GLOBALE'
    """)
    scorte_bt = {r["materiale"]: r["qty"] for r in cur.fetchall()}

    giorni_periodo = max((filtro_a - filtro_da).days, 1)
    proiezioni = {}
    for mat in BOTTIGLIE_GLOBALI:
        consumo_gg = consumo_bt.get(mat, 0) / giorni_periodo
        scorta     = scorte_bt.get(mat, 0)
        if consumo_gg > 0:
            proiezioni[mat] = round(scorta / consumo_gg)
        else:
            proiezioni[mat] = None

    cur.close()
    conn.close()

    return render_template("analisi.html",
        prod_cliente=prod_cliente,
        cons_cliente=cons_cliente,
        prod_prodotto=prod_prodotto,
        trend_prod=trend_prod,
        trend_cons=trend_cons,
        proiezioni=proiezioni,
        scorte_bt=scorte_bt,
        consumo_bt=consumo_bt,
        periodo=periodo,
        data_da=data_da,
        data_a=data_a,
        giorni_periodo=giorni_periodo,
    )


# ==================================================
# CALENDARIO
# ==================================================

@app.route("/calendario/mese")
def calendario_mese():
    from datetime import datetime, timedelta, date
    oggi = datetime.now().date()
    offset = int(request.args.get("offset", "0"))

    # Primo giorno del mese corrente + offset mesi
    anno = oggi.year
    mese = oggi.month + offset
    while mese > 12: mese -= 12; anno += 1
    while mese < 1:  mese += 12; anno -= 1

    primo_giorno = date(anno, mese, 1)
    if mese == 12:
        ultimo_giorno = date(anno+1, 1, 1) - timedelta(days=1)
    else:
        ultimo_giorno = date(anno, mese+1, 1) - timedelta(days=1)

    conn = db()
    cur = conn.cursor()
    cur.execute("""
        SELECT * FROM calendario_eventi
        WHERE (ricorrenza='nessuna' AND data_evento >= %s AND data_evento <= %s)
           OR ricorrenza IN ('settimanale','bisettimanale')
        ORDER BY data_evento
    """, (primo_giorno, ultimo_giorno))
    tutti_eventi = cur.fetchall()

    # Costruisce dict giorno -> lista eventi
    eventi_mese = {}
    d = primo_giorno
    while d <= ultimo_giorno:
        eventi_mese[d.isoformat()] = []
        d += timedelta(days=1)

    for e in tutti_eventi:
        if e["ricorrenza"] == "nessuna":
            key = e["data_evento"].isoformat()
            if key in eventi_mese:
                eventi_mese[key].append(dict(e))
        else:
            d = primo_giorno
            while d <= ultimo_giorno:
                if d.weekday() == e["data_evento"].weekday() and d >= e["data_evento"]:
                    if e["ricorrenza"] == "bisettimanale":
                        diff = (d - e["data_evento"]).days // 7
                        if diff % 2 != 0:
                            d += timedelta(days=1); continue
                    cur.execute("""
                        SELECT id FROM calendario_eccezioni
                        WHERE evento_id=%s AND data_eccezione=%s
                    """, (e["id"], d))
                    if not cur.fetchone():
                        ev = dict(e); ev["data_evento"] = d
                        eventi_mese[d.isoformat()].append(ev)
                d += timedelta(days=1)

    cur.close()
    conn.close()

    # Costruisce la griglia del mese (settimane)
    # Inizia dal lunedì della settimana del primo giorno
    start = primo_giorno - timedelta(days=primo_giorno.weekday())
    settimane = []
    cur_d = start
    while cur_d <= ultimo_giorno or len(settimane) < 5:
        settimana = []
        for _ in range(7):
            settimana.append(cur_d)
            cur_d += timedelta(days=1)
        settimane.append(settimana)
        if cur_d > ultimo_giorno and len(settimane) >= 4:
            break

    from datetime import timedelta as td
    return render_template("calendario_mese.html",
        settimane=settimane,
        eventi_mese=eventi_mese,
        primo_giorno=primo_giorno,
        ultimo_giorno=ultimo_giorno,
        oggi=oggi,
        offset=offset,
        mese_nome=["","Gennaio","Febbraio","Marzo","Aprile","Maggio","Giugno",
                   "Luglio","Agosto","Settembre","Ottobre","Novembre","Dicembre"][mese],
        anno=anno,
        mese=mese,
        timedelta=td,
    )


@app.route("/calendario")
def calendario():
    from datetime import datetime, timedelta
    oggi = datetime.now().date()
    # Lunedi della settimana corrente
    offset = request.args.get("offset", "0")
    try:
        offset = int(offset)
    except:
        offset = 0
    lun = oggi - timedelta(days=oggi.weekday()) + timedelta(weeks=offset)
    giorni = [lun + timedelta(days=i) for i in range(7)]

    conn = db()
    cur = conn.cursor()

    # Carica eventi della settimana (incluse ricorrenze)
    cur.execute("""
        SELECT * FROM calendario_eventi
        WHERE (
            (ricorrenza = 'nessuna' AND data_evento >= %s AND data_evento <= %s)
            OR ricorrenza IN ('settimanale', 'bisettimanale')
        )
        ORDER BY data_evento, ora_inizio
    """, (lun, lun + timedelta(days=6)))
    tutti_eventi = cur.fetchall()

    # Filtra ricorrenze per questa settimana
    eventi_settimana = {}
    for g in giorni:
        eventi_settimana[g.isoformat()] = []

    for e in tutti_eventi:
        if e["ricorrenza"] == "nessuna":
            key = e["data_evento"].isoformat()
            if key in eventi_settimana:
                eventi_settimana[key].append(dict(e))
        elif e["ricorrenza"] == "settimanale":
            dow_evento = e["data_evento"].weekday()
            for g in giorni:
                if g.weekday() == dow_evento and g >= e["data_evento"]:
                    # Controlla se non è stato eliminato per questa data
                    cur.execute("""
                        SELECT id FROM calendario_eccezioni
                        WHERE evento_id = %s AND data_eccezione = %s
                    """, (e["id"], g))
                    if not cur.fetchone():
                        ev = dict(e)
                        ev["data_evento"] = g
                        eventi_settimana[g.isoformat()].append(ev)
        elif e["ricorrenza"] == "bisettimanale":
            dow_evento = e["data_evento"].weekday()
            delta = (lun - e["data_evento"]).days
            settimane_passate = delta // 7
            for g in giorni:
                if g.weekday() == dow_evento and g >= e["data_evento"]:
                    diff = (g - e["data_evento"]).days // 7
                    if diff % 2 == 0:
                        cur.execute("""
                            SELECT id FROM calendario_eccezioni
                            WHERE evento_id = %s AND data_eccezione = %s
                        """, (e["id"], g))
                        if not cur.fetchone():
                            ev = dict(e)
                            ev["data_evento"] = g
                            eventi_settimana[g.isoformat()].append(ev)

    # Appunti weekend
    cur.execute("""
        SELECT * FROM calendario_note_weekend
        WHERE settimana_inizio = %s
    """, (lun,))
    nota_weekend = cur.fetchone()

    cur.close()
    conn.close()

    # Frase motivazionale della settimana
    frasi = [
        "Chi lavora con passione non conta le ore.",
        "Ogni bottiglia racconta una storia di cura e dedizione.",
        "La qualità non è mai un caso, è sempre il risultato di uno sforzo intelligente.",
        "Il vino è poesia in bottiglia.",
        "Lavorare bene oggi è il miglior investimento per domani.",
        "La costanza batte il talento quando il talento non si impegna.",
        "Ogni giorno è una nuova occasione per fare meglio.",
        "Il successo non arriva per caso, si costruisce giorno per giorno.",
        "Chi non si ferma vince.",
        "La qualità si ricorda molto dopo che il prezzo è stato dimenticato.",
    ]
    from datetime import date
    num_settimana = lun.isocalendar()[1]
    frase = frasi[num_settimana % len(frasi)]

    from datetime import timedelta as td
    return render_template("calendario.html",
        giorni=giorni,
        eventi=eventi_settimana,
        offset=offset,
        oggi=oggi,
        nota_weekend=nota_weekend,
        frase=frase,
        lun=lun,
        timedelta=td,
    )


@app.route("/aggiungi_evento", methods=["POST"])
def aggiungi_evento():
    from datetime import datetime
    data_ev  = request.form["data_evento"]
    titolo   = request.form["titolo"]
    categoria= request.form.get("categoria", "Altro")
    ricorr   = request.form.get("ricorrenza", "nessuna")
    ora_ini  = request.form.get("ora_inizio", "08:00")
    note     = request.form.get("note", "")
    offset   = request.form.get("offset", "0")

    conn = db()
    cur = conn.cursor()
    cur.execute("""
        INSERT INTO calendario_eventi
        (data_evento, titolo, categoria, ricorrenza, ora_inizio, note)
        VALUES (%s,%s,%s,%s,%s,%s)
    """, (data_ev, titolo, categoria, ricorr, ora_ini, note))
    conn.commit()
    cur.close()
    conn.close()
    return redirect(f"/calendario?offset={offset}")


@app.route("/elimina_evento", methods=["POST"])
def elimina_evento():
    evento_id  = request.form["evento_id"]
    tipo       = request.form["tipo"]  # 'questa' o 'tutte'
    data_ev    = request.form["data_evento"]
    offset     = request.form.get("offset", "0")

    conn = db()
    cur = conn.cursor()

    if tipo == "tutte":
        cur.execute("DELETE FROM calendario_eventi WHERE id=%s", (evento_id,))
        cur.execute("DELETE FROM calendario_eccezioni WHERE evento_id=%s", (evento_id,))
    else:
        # Aggiunge eccezione per questa data
        cur.execute("""
            INSERT INTO calendario_eccezioni (evento_id, data_eccezione)
            VALUES (%s,%s) ON CONFLICT DO NOTHING
        """, (evento_id, data_ev))

    conn.commit()
    cur.close()
    conn.close()
    return redirect(f"/calendario?offset={offset}")


@app.route("/salva_nota_weekend", methods=["POST"])
def salva_nota_weekend():
    lun   = request.form["settimana_inizio"]
    testo = request.form["testo"]
    offset= request.form.get("offset", "0")
    conn = db()
    cur = conn.cursor()
    cur.execute("""
        INSERT INTO calendario_note_weekend (settimana_inizio, testo)
        VALUES (%s,%s)
        ON CONFLICT (settimana_inizio) DO UPDATE SET testo=EXCLUDED.testo
    """, (lun, testo))
    conn.commit()
    cur.close()
    conn.close()
    return redirect(f"/calendario?offset={offset}")


# ==================================================
# CRONOMETRO
# ==================================================

@app.route("/cronometro")
def cronometro():
    conn = db()
    cur = conn.cursor()
    cur.execute("SELECT * FROM produzione WHERE done=0 ORDER BY cliente, prodotto")
    ordini = cur.fetchall()
    cur.execute("""
        SELECT * FROM sessioni_cronometro
        ORDER BY data_inizio DESC LIMIT 20
    """)
    sessioni = cur.fetchall()
    cur.close()
    conn.close()
    return render_template("cronometro.html",
        ordini=ordini,
        sessioni=sessioni,
        moltiplicatori=MOLTIPLICATORI,
        pedane_config=PEDANE_CONFIG,
    )


@app.route("/start_sessione", methods=["POST"])
def start_sessione():
    from datetime import datetime
    import json as _json

    conn = db()
    cur = conn.cursor()

    # Prende tutti gli ordini in produzione non completati
    cur.execute("SELECT * FROM produzione WHERE done=0 ORDER BY id")
    ordini = cur.fetchall()

    dettagli = []
    tempo_prod_min = 0
    prodotti_unici = set()
    formati = set()

    for o in ordini:
        molt = MOLTIPLICATORI.get(o["cliente"], {}).get(o["prodotto"], 1)
        fardelli = o["qty"] / molt
        formato = "2L" if "2L" in o["prodotto"] else "1L"
        fard_ped = PEDANE_CONFIG.get(o["cliente"], {}).get(formato, 64)
        if fard_ped:
            pedane = fardelli / fard_ped
            tempo_prod_min += pedane * get_tempo_pedana(o["cliente"], formato)
        prodotti_unici.add(o["prodotto"])
        formati.add(formato)
        dettagli.append({
            "id": o["id"],
            "cliente": o["cliente"],
            "prodotto": o["prodotto"],
            "qty": o["qty"],
            "fardelli": round(fardelli, 1),
        })

    # Stima con tutte le voci
    tempo_stimato = tempo_prod_min
    tempo_stimato += 30 + 60                  # riscaldamento + pranzo
    tempo_stimato += len(prodotti_unici) * 3  # cambio etichetta 3 min
    if len(formati) > 1:
        tempo_stimato += 10                   # cambio formato
    tempo_stimato = round(tempo_stimato * 1.075)

    cur.execute("""
        INSERT INTO sessioni_cronometro
        (data_inizio, ordini_json, cambi_formato, cambi_etichetta, tempo_stimato_min, stato)
        VALUES (%s,%s,%s,%s,%s,'in_corso')
        RETURNING id
    """, (
        now_rome(),
        _json.dumps(dettagli),
        1 if len(formati) > 1 else 0,
        len(prodotti_unici),
        tempo_stimato,
    ))
    conn.commit()
    cur.close()
    conn.close()
    return redirect("/produzione")


@app.route("/stop_sessione/<int:sid>", methods=["POST"])
def stop_sessione(sid):
    from datetime import datetime
    conn = db()
    cur = conn.cursor()
    cur.execute("SELECT * FROM sessioni_cronometro WHERE id=%s", (sid,))
    sess = cur.fetchone()
    if sess:
        ora_fine = now_rome()
        minuti_reali = round((ora_fine - sess["data_inizio"]).total_seconds() / 60)
        cur.execute("""
            UPDATE sessioni_cronometro
            SET data_fine=%s, tempo_reale_min=%s, stato='completata'
            WHERE id=%s
        """, (ora_fine, minuti_reali, sid))
        conn.commit()
    cur.close()
    conn.close()
    return redirect("/produzione")


# ==================================================
# BACKUP E RIPRISTINO
# ==================================================
@app.route("/backup_manuale")
def backup_manuale():
    import json as _json
    from datetime import datetime as _dt

    conn = db()
    cur = conn.cursor()
    tabelle = ["stock", "produzione", "storico", "note", "materie_prime", "storico_mp", "calendario_eventi", "calendario_eccezioni", "calendario_note_weekend", "sessioni_cronometro"]
    backup = {"data_backup": _dt.now().isoformat(), "tabelle": {}}

    for tabella in tabelle:
        cur.execute(f"SELECT * FROM {tabella} ORDER BY id")
        righe = []
        for r in cur.fetchall():
            riga = dict(r)
            for k, v in riga.items():
                if hasattr(v, "isoformat"):
                    riga[k] = v.isoformat()
            righe.append(riga)
        backup["tabelle"][tabella] = righe

    cur.close()
    conn.close()

    nome = f"backup_{_dt.now().strftime('%Y%m%d_%H%M%S')}.json"
    path = os.path.join(BASE_DIR, nome)
    with open(path, "w", encoding="utf-8") as f:
        _json.dump(backup, f, ensure_ascii=False, indent=2)

    return send_file(path, as_attachment=True, download_name=nome)


@app.route("/ripristino", methods=["GET", "POST"])
def ripristino():
    """
    Pagina per caricare un file JSON di backup e ripristinare il database.
    """
    from flask import flash
    import json as _json

    if request.method == "GET":
        return render_template("ripristino.html")

    file = request.files.get("backup_file")
    if not file or not file.filename.endswith(".json"):
        return render_template("ripristino.html", errore="Seleziona un file .json valido")

    try:
        data = _json.load(file)
        tabelle = data.get("tabelle", {})

        conn = db()
        cur = conn.cursor()

        ripristinate = 0
        righe_totali = 0

        for tabella, righe in tabelle.items():
            if tabella not in ["stock","produzione","storico","note","materie_prime","storico_mp","calendario_eventi","calendario_eccezioni","calendario_note_weekend","sessioni_cronometro"]:
                continue

            # Svuota la tabella e reimposta la sequenza
            cur.execute(f"DELETE FROM {tabella}")
            cur.execute(f"ALTER SEQUENCE {tabella}_id_seq RESTART WITH 1")

            for r in righe:
                cols = list(r.keys())
                vals = list(r.values())
                placeholders = ",".join(["%s"] * len(cols))
                col_str = ",".join(cols)
                cur.execute(
                    f"INSERT INTO {tabella}({col_str}) VALUES({placeholders})",
                    vals
                )
                righe_totali += 1

            # Riallinea la sequenza dopo l'import
            cur.execute(f"""
                SELECT setval('{tabella}_id_seq',
                    COALESCE((SELECT MAX(id) FROM {tabella}), 1))
            """)
            ripristinate += 1

        conn.commit()
        cur.close()
        conn.close()

        return render_template("ripristino.html",
            successo=f"Ripristino completato: {ripristinate} tabelle, {righe_totali} righe importate.")

    except Exception as e:
        return render_template("ripristino.html", errore=f"Errore durante il ripristino: {str(e)}")


# ==================================================
# START
# ==================================================
if __name__ == "__main__":
    app.run(host="0.0.0.0", port=10000)
